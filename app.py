import os
import io
import csv
import json
import hashlib
import sqlite3
import logging
from datetime import datetime
from functools import wraps
from threading import Lock

from flask import Flask, render_template, jsonify, request, g, Response
from flask_socketio import SocketIO, emit
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

# --- Logging ---
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv("FLASK_SECRET", "change-me")
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB max upload
socketio = SocketIO(app, cors_allowed_origins="*", async_mode="eventlet")

# --- Autenticação básica ---
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "")

def check_auth(password):
    """Verifica se a senha está correta."""
    return password == ADMIN_PASSWORD

def requires_auth(f):
    """Decorator para proteger rotas com senha (se ADMIN_PASSWORD estiver definida)."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not ADMIN_PASSWORD:
            return f(*args, **kwargs)  # sem senha configurada, acesso livre
        auth = request.authorization
        if not auth or not check_auth(auth.password):
            return Response(
                'Acesso negado. Informe a senha de administrador.',
                401,
                {'WWW-Authenticate': 'Basic realm="Check-in Universal"'}
            )
        return f(*args, **kwargs)
    return decorated

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
PARTICIPANTS_FILE = os.path.join(DATA_DIR, "participants.json")
CONFIG_FILE = os.path.join(DATA_DIR, "config.json")
DATABASE = os.path.join(DATA_DIR, "checkins.db")

os.makedirs(DATA_DIR, exist_ok=True)

# --- Banco de dados SQLite ---

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
        db.row_factory = sqlite3.Row
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def init_db():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS checkins (
            external_id TEXT PRIMARY KEY,
            checked_in INTEGER NOT NULL,
            checked_by TEXT,
            checked_at TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

init_db()

# --- Helpers para JSON ---

def load_json(path, default=None):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return default if default is not None else {}

def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def generate_id(col1_val, col2_val):
    """Gera um ID determinístico baseado nos valores das duas colunas."""
    raw = f"{col1_val or ''}|{col2_val or ''}".strip().lower()
    return hashlib.md5(raw.encode()).hexdigest()[:12]

# --- Checkins (SQLite) ---

def load_checkins_from_db():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute('SELECT external_id, checked_in, checked_by, checked_at FROM checkins')
    rows = cursor.fetchall()
    conn.close()
    checkins = {}
    for row in rows:
        checkins[row[0]] = {
            "checked_in": bool(row[1]),
            "checked_by": row[2],
            "checked_at": row[3]
        }
    return checkins

def save_checkin_to_db(external_id, checked_in, checked_by, checked_at):
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO checkins (external_id, checked_in, checked_by, checked_at, updated_at)
        VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(external_id) DO UPDATE SET
            checked_in = excluded.checked_in,
            checked_by = excluded.checked_by,
            checked_at = excluded.checked_at,
            updated_at = CURRENT_TIMESTAMP
    ''', (external_id, 1 if checked_in else 0, checked_by, checked_at))
    conn.commit()
    conn.close()

# --- Participantes ---

participants_cache = []
cache_lock = Lock()

def load_participants():
    """Carrega participantes do JSON e mescla com estado de checkin do banco."""
    raw = load_json(PARTICIPANTS_FILE, [])
    checkins_db = load_checkins_from_db()
    with cache_lock:
        global participants_cache
        participants_cache = []
        for p in raw:
            ext = p.get("external_id", "")
            state = checkins_db.get(ext, {})
            p["_checked_in"] = bool(state.get("checked_in", False))
            p["_checked_by"] = state.get("checked_by")
            p["_checked_at"] = state.get("checked_at")
            participants_cache.append(p)
    return participants_cache

# Carrega ao iniciar se já existir
if os.path.exists(PARTICIPANTS_FILE):
    load_participants()

# --- Rotas HTTP ---

@app.route("/")
@requires_auth
def index():
    config = load_json(CONFIG_FILE, {})
    return render_template("index.html", config=config)

@app.route("/api/config", methods=["GET"])
def api_get_config():
    """Retorna a configuração atual (nomes das colunas)."""
    config = load_json(CONFIG_FILE, {})
    return jsonify(config)

@app.route("/api/upload", methods=["POST"])
def api_upload():
    """
    Recebe um arquivo XLSX + nomes das duas colunas.
    Lê a planilha, gera IDs e salva em participants.json.
    """
    if 'file' not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "Formato inválido. Envie um arquivo .xlsx"}), 400

    col1_name = request.form.get("col1_name", "").strip()
    col2_name = request.form.get("col2_name", "").strip()
    has_qr = request.form.get("has_qr", "false").lower() == "true"
    qr_col_name = request.form.get("qr_col_name", "").strip()

    if not col1_name or not col2_name:
        return jsonify({"error": "Informe os nomes das duas colunas"}), 400

    if has_qr and not qr_col_name:
        return jsonify({"error": "Informe o nome da coluna do QR Code"}), 400

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active

        # Ler cabeçalhos da primeira linha
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        headers_lower = [str(h).strip().lower() if h else "" for h in headers]

        col1_idx = None
        col2_idx = None
        qr_col_idx = None
        for i, h in enumerate(headers_lower):
            if h == col1_name.lower():
                col1_idx = i
            if h == col2_name.lower():
                col2_idx = i
            if has_qr and qr_col_name and h == qr_col_name.lower():
                qr_col_idx = i

        if col1_idx is None:
            return jsonify({"error": f"Coluna '{col1_name}' não encontrada. Colunas disponíveis: {[h for h in headers if h]}"}), 400
        if col2_idx is None:
            return jsonify({"error": f"Coluna '{col2_name}' não encontrada. Colunas disponíveis: {[h for h in headers if h]}"}), 400
        if has_qr and qr_col_idx is None:
            return jsonify({"error": f"Coluna QR '{qr_col_name}' não encontrada. Colunas disponíveis: {[h for h in headers if h]}"}), 400

        participants = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            val1 = str(row[col1_idx]).strip() if row[col1_idx] is not None else ""
            val2 = str(row[col2_idx]).strip() if row[col2_idx] is not None else ""
            if not val1 and not val2:
                continue  # pular linhas vazias
            ext_id = generate_id(val1, val2)
            entry = {
                "external_id": ext_id,
                "col1": val1,
                "col2": val2
            }
            if has_qr and qr_col_idx is not None:
                qr_val = str(row[qr_col_idx]).strip() if row[qr_col_idx] is not None else ""
                entry["qr_code"] = qr_val
            participants.append(entry)
        wb.close()
    except Exception as e:
        return jsonify({"error": f"Erro ao processar planilha: {str(e)}"}), 500

    # Salvar configuração
    config = {
        "col1_name": col1_name,
        "col2_name": col2_name,
        "has_qr": has_qr,
        "qr_col_name": qr_col_name if has_qr else "",
        "filename": file.filename,
        "total": len(participants),
        "uploaded_at": datetime.now().isoformat()
    }
    save_json(CONFIG_FILE, config)

    # Salvar participantes
    save_json(PARTICIPANTS_FILE, participants)

    # Recarregar cache
    load_participants()
    logger.info(f"Planilha '{file.filename}' processada: {len(participants)} participantes")

    # Notificar todos os clientes conectados
    socketio.emit("participants_updated", {"total": len(participants)})

    return jsonify({
        "ok": True,
        "total": len(participants),
        "config": config
    })

@app.route("/api/participants", methods=["GET"])
def api_participants():
    """Retorna lista de participantes com estado de checkin."""
    config = load_json(CONFIG_FILE, {})

    if not participants_cache:
        load_participants()

    q = request.args.get("q", "").strip().lower()
    if q:
        filtered = [
            p for p in participants_cache
            if q in (p.get("col1", "").lower()) or q in (p.get("col2", "").lower()) or q in (p.get("qr_code", "").lower())
        ]
    else:
        filtered = list(participants_cache)

    return jsonify({"participants": filtered, "config": config})

@app.route("/api/checkin", methods=["POST"])
def api_checkin():
    """Marca checkin por external_id."""
    data = request.get_json() or {}
    external_id = data.get("external_id")
    fiscal = data.get("fiscal", "Unknown")

    if not external_id:
        return jsonify({"error": "external_id é obrigatório"}), 400

    now_iso = datetime.now().isoformat() + "Z"
    save_checkin_to_db(external_id, True, fiscal, now_iso)

    with cache_lock:
        for p in participants_cache:
            if p.get("external_id") == external_id:
                p["_checked_in"] = True
                p["_checked_by"] = fiscal
                p["_checked_at"] = now_iso

    payload = {
        "external_id": external_id,
        "checked_in": True,
        "checked_by": fiscal,
        "checked_at": now_iso
    }
    socketio.emit("checkin_update", payload)
    return jsonify({"ok": True, "payload": payload})

@app.route("/api/uncheck", methods=["POST"])
def api_uncheck():
    """Desfaz checkin por external_id."""
    data = request.get_json() or {}
    external_id = data.get("external_id")
    if not external_id:
        return jsonify({"error": "external_id é obrigatório"}), 400

    save_checkin_to_db(external_id, False, None, None)

    with cache_lock:
        for p in participants_cache:
            if p.get("external_id") == external_id:
                p["_checked_in"] = False
                p["_checked_by"] = None
                p["_checked_at"] = None

    payload = {"external_id": external_id, "checked_in": False}
    socketio.emit("checkin_update", payload)
    return jsonify({"ok": True, "payload": payload})

@app.route("/api/checkins/export", methods=["GET"])
def export_checkins():
    """Exporta checkins como CSV ou JSON."""
    fmt = request.args.get("format", "json").lower()
    checkins_db = load_checkins_from_db()
    config = load_json(CONFIG_FILE, {})
    participants = load_json(PARTICIPANTS_FILE, [])

    if fmt == "csv":
        col1_name = config.get("col1_name", "Coluna 1")
        col2_name = config.get("col2_name", "Coluna 2")

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([col1_name, col2_name, "Check-in", "Fiscal", "Data/Hora"])

        for p in participants:
            ext = p.get("external_id", "")
            state = checkins_db.get(ext, {})
            writer.writerow([
                p.get("col1", ""),
                p.get("col2", ""),
                "Sim" if state.get("checked_in") else "Não",
                state.get("checked_by", ""),
                state.get("checked_at", "")
            ])

        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment; filename=checkins_export.csv'}
        )

    # JSON padrão
    return jsonify({"checkins": checkins_db, "total": len(checkins_db)})

@app.route("/api/reset", methods=["POST"])
def api_reset():
    """Reseta todos os checkins (limpa o banco)."""
    conn = sqlite3.connect(DATABASE)
    conn.execute("DELETE FROM checkins")
    conn.commit()
    conn.close()
    load_participants()
    socketio.emit("participants_updated", {"reset": True})
    return jsonify({"ok": True})

# --- SocketIO ---

@socketio.on("connect")
def on_connect():
    checkins = load_checkins_from_db()
    emit("initial_state", {"checkin_state": checkins})

if __name__ == "__main__":
    debug_mode = os.getenv("FLASK_DEBUG", "false").lower() in ("true", "1", "yes")
    port = int(os.getenv("PORT", "5030"))
    logger.info(f"Iniciando servidor na porta {port} (debug={debug_mode})")
    if ADMIN_PASSWORD:
        logger.info("Autenticação por senha ATIVADA")
    else:
        logger.warning("ADMIN_PASSWORD não definida — acesso sem senha!")
    socketio.run(app, host="0.0.0.0", port=port, debug=debug_mode)
